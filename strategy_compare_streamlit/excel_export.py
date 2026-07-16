"""
Excel Export Engine
"""

from datetime import datetime

from pathlib import Path

import json

import pandas as pd

from openpyxl.styles import (

    Font,

    PatternFill,

    Border,

    Side,

    Alignment

)

from openpyxl.utils import get_column_letter

from config import (

    REPORT_NAME,

    REPORTS

)


###########################################################################
# EXCEL EXPORTER
###########################################################################

class ExcelExporter:

    def __init__(

        self,

        output=REPORT_NAME

    ):

        self.output = Path(

            output

        )

        self.output.parent.mkdir(

            parents=True,

            exist_ok=True

        )

        #######################################################################
        # FORMATTING
        #######################################################################

        self.header_fill = PatternFill(

            fill_type="solid",

            fgColor="1F4E78"

        )

        self.header_font = Font(

            bold=True,

            color="FFFFFF"

        )

        self.border = Border(

            left=Side(

                style="thin"

            ),

            right=Side(

                style="thin"

            ),

            top=Side(

                style="thin"

            ),

            bottom=Side(

                style="thin"

            )

        )

        self.center = Alignment(

            horizontal="center",

            vertical="center"

        )

    ###########################################################################
    # FORMAT SHEET
    ###########################################################################

    def format_sheet(

        self,

        worksheet

    ):

        worksheet.freeze_panes = "A2"

        worksheet.auto_filter.ref = (

            worksheet.dimensions

        )

        for cell in worksheet[1]:

            cell.fill = self.header_fill

            cell.font = self.header_font

            cell.border = self.border

            cell.alignment = self.center

        self.auto_width(

            worksheet

        )

    ###########################################################################
    # AUTO COLUMN WIDTH
    ###########################################################################

    def auto_width(

        self,

        worksheet

    ):

        for column in worksheet.columns:

            length = 0

            letter = get_column_letter(

                column[0].column

            )

            for cell in column:

                try:

                    value = str(

                        cell.value

                    )

                    if len(

                        value

                    ) > length:

                        length = len(

                            value

                        )

                except Exception:

                    pass

            worksheet.column_dimensions[

                letter

            ].width = min(

                max(

                    length + 3,

                    12

                ),

                40

            )

    ###########################################################################
    # WRITE DATAFRAME
    ###########################################################################

    def write_sheet(

        self,

        writer,

        dataframe,

        sheet_name

    ):

        dataframe.to_excel(

            writer,

            sheet_name=sheet_name,

            index=False

        )

        worksheet = writer.sheets[

            sheet_name

        ]

        self.format_sheet(

            worksheet

        )

    ###########################################################################
    # EXPORT WORKBOOK
    ###########################################################################

    def export(

        self,

        ranked,

        recommendations,

        overlap

    ):

        summary = pd.DataFrame({

            "Property": [

                "Generated On",

                "Total Strategies",

                "Best Strategy",

                "Best Overall Score",

                "Average Overall Score",

                "Strong Buy",

                "Buy",

                "Watch",

                "Improve",

                "Avoid",

                "Reject"

            ],

            "Value": [

                datetime.now().strftime(

                    "%Y-%m-%d %H:%M:%S"

                ),

                len(

                    ranked

                ),

                ranked.iloc[0][

                    "Strategy"

                ]

                if not ranked.empty

                else "",

                round(

                    ranked[

                        "Overall Score"

                    ].max(),

                    2

                )

                if not ranked.empty

                else "",

                round(

                    ranked[

                        "Overall Score"

                    ].mean(),

                    2

                )

                if not ranked.empty

                else "",

                int(

                    (

                        ranked[

                            "Recommendation"

                        ]

                        ==

                        "Strong Buy"

                    ).sum()

                ),

                int(

                    (

                        ranked[

                            "Recommendation"

                        ]

                        ==

                        "Buy"

                    ).sum()

                ),

                int(

                    (

                        ranked[

                            "Recommendation"

                        ]

                        ==

                        "Watch"

                    ).sum()

                ),

                int(

                    (

                        ranked[

                            "Recommendation"

                        ]

                        ==

                        "Improve"

                    ).sum()

                ),

                int(

                    (

                        ranked[

                            "Recommendation"

                        ]

                        ==

                        "Avoid"

                    ).sum()

                ),

                int(

                    (

                        ranked[

                            "Recommendation"

                        ]

                        ==

                        "Reject"

                    ).sum()

                )

            ]

        })

        metadata = pd.DataFrame({

            "Attribute": [

                "Application",

                "Version",

                "Output Folder",

                "Workbook",

                "Generated By"

            ],

            "Value": [

                "Strategy Comparison Dashboard",

                "V3",

                str(

                    REPORTS

                ),

                self.output.name,

                "ExcelExporter"

            ]

        })

        with pd.ExcelWriter(

            self.output,

            engine="openpyxl"

        ) as writer:

            ###################################################################
            # SUMMARY
            ###################################################################

            self.write_sheet(

                writer,

                summary,

                "Executive Summary"

            )

            ###################################################################
            # RANKING
            ###################################################################

            self.write_sheet(

                writer,

                ranked,

                "Strategy Ranking"

            )

            ###################################################################
            # RECOMMENDATIONS
            ###################################################################

            self.write_sheet(

                writer,

                recommendations,

                "Recommendations"

            )

            ###################################################################
            # OVERLAP
            ###################################################################

            self.write_sheet(

                writer,

                overlap,

                "Top Stock Overlap"

            )

            ###################################################################
            # METADATA
            ###################################################################

            self.write_sheet(

                writer,

                metadata,

                "Metadata"

            )

        return self.output
    
    ###########################################################################
    # EXPORT CSV
    ###########################################################################

    def export_csv(

        self,

        dataframe,

        filename

    ):

        path = Path(

            filename

        )

        path.parent.mkdir(

            parents=True,

            exist_ok=True

        )

        dataframe.to_csv(

            path,

            index=False,

            encoding="utf-8"

        )

        return path

    ###########################################################################
    # EXPORT JSON
    ###########################################################################

    def export_json(

        self,

        dataframe,

        filename

    ):

        path = Path(

            filename

        )

        path.parent.mkdir(

            parents=True,

            exist_ok=True

        )

        dataframe.to_json(

            path,

            orient="records",

            indent=4

        )

        return path

    ###########################################################################
    # EXPORT MULTIPLE DATASETS
    ###########################################################################

    def export_multiple(

        self,

        datasets

    ):

        exported = {}

        for name, dataframe in datasets.items():

            csv_file = (

                REPORTS

                /

                f"{name}.csv"

            )

            json_file = (

                REPORTS

                /

                f"{name}.json"

            )

            exported[

                f"{name}_csv"

            ] = self.export_csv(

                dataframe,

                csv_file

            )

            exported[

                f"{name}_json"

            ] = self.export_json(

                dataframe,

                json_file

            )

        return exported

    ###########################################################################
    # EXPORT VALIDATION
    ###########################################################################

    def validate(

        self

    ):

        return {

            "Workbook Exists":

                self.output.exists(),

            "Workbook":

                str(

                    self.output

                )

        }

    ###########################################################################
    # EXPORT REPORT
    ###########################################################################

    def report(

        self,

        ranked,

        recommendations,

        overlap

    ):

        workbook = self.export(

            ranked,

            recommendations,

            overlap

        )

        exports = self.export_multiple({

            "Strategy_Ranking":

                ranked,

            "Recommendations":

                recommendations,

            "Top_Stock_Overlap":

                overlap

        })

        return {

            "excel":

                workbook,

            "csv_json":

                exports,

            "validation":

                self.validate()

        }