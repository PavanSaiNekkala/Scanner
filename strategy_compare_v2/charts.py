from openpyxl.chart import BarChart
from openpyxl.chart import Reference
from openpyxl.styles import Font
from openpyxl.chart import ScatterChart
from openpyxl.chart import Series
from openpyxl.chart import PieChart
from collections import Counter


class ChartBuilder:

    def __init__(self, workbook):

        self.wb = workbook

    ###########################################################################
    # OVERALL SCORE CHART
    ###########################################################################

    def overall_score_chart(self):

        if "Strategy Ranking" not in self.wb.sheetnames:
            return

        ws = self.wb["Strategy Ranking"]

        chart = BarChart()

        chart.type = "bar"

        chart.style = 10

        chart.title = "Overall Strategy Score"

        chart.y_axis.title = "Strategy"

        chart.x_axis.title = "Score"

        score_col = None

        strategy_col = None

        for cell in ws[1]:

            if cell.value == "Overall Score":
                score_col = cell.column

            if cell.value == "Strategy":
                strategy_col = cell.column

        if score_col is None or strategy_col is None:
            return

        data = Reference(
            ws,
            min_col=score_col,
            min_row=1,
            max_row=ws.max_row
        )

        cats = Reference(
            ws,
            min_col=strategy_col,
            min_row=2,
            max_row=ws.max_row
        )

        chart.add_data(
            data,
            titles_from_data=True
        )

        chart.set_categories(cats)

        chart.height = 8

        chart.width = 18

        ws.add_chart(
            chart,
            "L2"
        )

    ###########################################################################
    # COMPOSITE SCORE CHART
    ###########################################################################

    def composite_score_chart(self):

        if "Strategy Ranking" not in self.wb.sheetnames:
            return

        ws = self.wb["Strategy Ranking"]

        composite_col = None

        strategy_col = None

        for cell in ws[1]:

            if cell.value == "Composite Score_Mean":
                composite_col = cell.column

            if cell.value == "Strategy":
                strategy_col = cell.column

        if composite_col is None:
            return

        chart = BarChart()

        chart.type = "col"

        chart.style = 11

        chart.title = "Average Composite Score"

        chart.y_axis.title = "Composite Score"

        data = Reference(
            ws,
            min_col=composite_col,
            min_row=1,
            max_row=ws.max_row
        )

        cats = Reference(
            ws,
            min_col=strategy_col,
            min_row=2,
            max_row=ws.max_row
        )

        chart.add_data(
            data,
            titles_from_data=True
        )

        chart.set_categories(cats)

        chart.height = 8

        chart.width = 16

        ws.add_chart(
            chart,
            "L20"
        )

    ###########################################################################
    # PROFIT FACTOR CHART
    ###########################################################################

    def profit_factor_chart(self):

        if "Strategy Ranking" not in self.wb.sheetnames:
            return

        ws = self.wb["Strategy Ranking"]

        pf_col = None

        strategy_col = None

        for cell in ws[1]:

            if cell.value == "Profit Factor_Mean":
                pf_col = cell.column

            if cell.value == "Strategy":
                strategy_col = cell.column

        if pf_col is None:
            return

        chart = BarChart()

        chart.style = 12

        chart.type = "col"

        chart.title = "Average Profit Factor"

        chart.y_axis.title = "Profit Factor"

        data = Reference(
            ws,
            min_col=pf_col,
            min_row=1,
            max_row=ws.max_row
        )

        cats = Reference(
            ws,
            min_col=strategy_col,
            min_row=2,
            max_row=ws.max_row
        )

        chart.add_data(
            data,
            titles_from_data=True
        )

        chart.set_categories(cats)

        chart.height = 8

        chart.width = 16

        ws.add_chart(
            chart,
            "L38"
        )

    ###########################################################################
    # BUILD ALL BAR CHARTS
    ###########################################################################

    def build(self):

        self.overall_score_chart()

        self.composite_score_chart()

        self.profit_factor_chart()

        self.recommendation_pie_chart()

        self.grade_pie_chart()

        self.analytics_charts()

    ###########################################################################
    # RECOMMENDATION PIE CHART
    ###########################################################################

    def recommendation_pie_chart(self):

        if "Dashboard" not in self.wb.sheetnames:
            return

        if "Strategy Ranking" not in self.wb.sheetnames:
            return

        dashboard = self.wb["Dashboard"]

        ranking = self.wb["Strategy Ranking"]

        recommendation_col = None

        for cell in ranking[1]:

            if cell.value == "Recommendation":

                recommendation_col = cell.column

                break

        if recommendation_col is None:
            return

        recommendations = []

        for row in range(2, ranking.max_row + 1):

            recommendations.append(

                ranking.cell(
                    row=row,
                    column=recommendation_col
                ).value

            )

        counts = Counter(recommendations)

        start_row = 30

        dashboard.cell(
            row=start_row,
            column=1
        ).value = "Recommendation"

        dashboard.cell(
            row=start_row,
            column=2
        ).value = "Count"

        current = start_row + 1

        for key, value in counts.items():

            dashboard.cell(
                row=current,
                column=1
            ).value = key

            dashboard.cell(
                row=current,
                column=2
            ).value = value

            current += 1

        chart = PieChart()

        chart.title = "Recommendation Distribution"

        labels = Reference(

            dashboard,

            min_col=1,

            min_row=start_row + 1,

            max_row=current - 1

        )

        data = Reference(

            dashboard,

            min_col=2,

            min_row=start_row,

            max_row=current - 1

        )

        chart.add_data(

            data,

            titles_from_data=True

        )

        chart.set_categories(labels)

        chart.height = 8

        chart.width = 10

        dashboard.add_chart(

            chart,

            "E30"

        )

    ###########################################################################
    # GRADE DISTRIBUTION PIE CHART
    ###########################################################################

    def grade_pie_chart(self):

        if "Dashboard" not in self.wb.sheetnames:
            return

        if "Strategy Ranking" not in self.wb.sheetnames:
            return

        dashboard = self.wb["Dashboard"]

        ranking = self.wb["Strategy Ranking"]

        grade_col = None

        for cell in ranking[1]:

            if cell.value == "Grade":

                grade_col = cell.column

                break

        if grade_col is None:
            return

        grades = []

        for row in range(2, ranking.max_row + 1):

            grades.append(

                ranking.cell(

                    row=row,

                    column=grade_col

                ).value

            )

        counts = Counter(grades)

        start_row = 30

        start_col = 9

        dashboard.cell(

            row=start_row,

            column=start_col

        ).value = "Grade"

        dashboard.cell(

            row=start_row,

            column=start_col + 1

        ).value = "Count"

        current = start_row + 1

        for key, value in counts.items():

            dashboard.cell(

                row=current,

                column=start_col

            ).value = key

            dashboard.cell(

                row=current,

                column=start_col + 1

            ).value = value

            current += 1

        chart = PieChart()

        chart.title = "Grade Distribution"

        labels = Reference(

            dashboard,

            min_col=start_col,

            min_row=start_row + 1,

            max_row=current - 1

        )

        data = Reference(

            dashboard,

            min_col=start_col + 1,

            min_row=start_row,

            max_row=current - 1

        )

        chart.add_data(

            data,

            titles_from_data=True

        )

        chart.set_categories(labels)

        chart.height = 8

        chart.width = 10

        dashboard.add_chart(

            chart,

            "M30"

        )

    ###########################################################################
    # RELIABILITY VS COMPOSITE SCORE
    ###########################################################################

    def reliability_vs_composite_chart(self):

        if "Strategy Ranking" not in self.wb.sheetnames:
            return

        ws = self.wb["Strategy Ranking"]

        reliability_col = None
        composite_col = None

        for cell in ws[1]:

            if cell.value == "Reliability Score_Mean":
                reliability_col = cell.column

            elif cell.value == "Composite Score_Mean":
                composite_col = cell.column

        if reliability_col is None or composite_col is None:
            return

        chart = ScatterChart()

        chart.title = "Reliability vs Composite Score"

        chart.x_axis.title = "Reliability"

        chart.y_axis.title = "Composite Score"

        xvalues = Reference(
            ws,
            min_col=reliability_col,
            min_row=2,
            max_row=ws.max_row
        )

        yvalues = Reference(
            ws,
            min_col=composite_col,
            min_row=2,
            max_row=ws.max_row
        )

        series = Series(
            yvalues,
            xvalues,
            title="Strategies"
        )

        chart.series.append(series)

        chart.height = 8

        chart.width = 12

        ws.add_chart(
            chart,
            "L56"
        )


    ###########################################################################
    # EXPECTANCY VS PROFIT FACTOR
    ###########################################################################

    def expectancy_vs_pf_chart(self):

        if "Strategy Ranking" not in self.wb.sheetnames:
            return

        ws = self.wb["Strategy Ranking"]

        expectancy_col = None
        pf_col = None

        for cell in ws[1]:

            if cell.value == "Expectancy%_Mean":
                expectancy_col = cell.column

            elif cell.value == "Profit Factor_Mean":
                pf_col = cell.column

        if expectancy_col is None or pf_col is None:
            return

        chart = ScatterChart()

        chart.title = "Expectancy vs Profit Factor"

        chart.x_axis.title = "Expectancy"

        chart.y_axis.title = "Profit Factor"

        xvalues = Reference(
            ws,
            min_col=expectancy_col,
            min_row=2,
            max_row=ws.max_row
        )

        yvalues = Reference(
            ws,
            min_col=pf_col,
            min_row=2,
            max_row=ws.max_row
        )

        series = Series(
            yvalues,
            xvalues,
            title="Strategies"
        )

        chart.series.append(series)

        chart.height = 8

        chart.width = 12

        ws.add_chart(
            chart,
            "L74"
        )

    ###########################################################################
    # BUILD ANALYTICS CHARTS
    ###########################################################################

    def analytics_charts(self):

        self.reliability_vs_composite_chart()

        self.expectancy_vs_pf_chart()