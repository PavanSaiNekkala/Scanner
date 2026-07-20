from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side


class DashboardSheet:
    def __init__(self, workbook):
        self.wb = workbook

    #######################################################################

    def create(self, ranked_df):
        ws = self.wb.create_sheet("Dashboard", 0)

        self.title(ws)

        self.summary_cards(ws, ranked_df)

        self.top5_table(ws, ranked_df)

        self.metric_summary(ws, ranked_df)

        return ws

    #######################################################################

    def title(self, ws):
        ws.merge_cells("A1:G2")

        cell = ws["A1"]

        cell.value = "Strategy Comparison Dashboard"

        cell.font = Font(size=18, bold=True, color="FFFFFF")

        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")

        cell.alignment = Alignment(horizontal="center", vertical="center")

    #######################################################################

    def summary_cards(self, ws, ranked_df):
        top = ranked_df.iloc[0]

        cards = [
            ("Best Strategy", top["Strategy"]),
            ("Overall Score", round(top["Overall Score"], 2)),
            ("Grade", top["Grade"]),
            ("Recommendation", top["Recommendation"]),
            ("Strategies Compared", len(ranked_df)),
        ]

        row = 4

        thin = Side(style="thin")

        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for title, value in cards:
            ws[f"A{row}"] = title

            ws[f"B{row}"] = value

            ws[f"A{row}"].font = Font(bold=True)

            ws[f"A{row}"].fill = PatternFill(fill_type="solid", fgColor="D9EAD3")

            ws[f"A{row}"].border = border

            ws[f"B{row}"].border = border

            row += 1

    #######################################################################

    def top5_table(self, ws, ranked_df):
        start = 12

        headers = ["Rank", "Strategy", "Score", "Grade"]

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=start, column=c)

            cell.value = h

            cell.font = Font(bold=True, color="FFFFFF")

            cell.fill = PatternFill(fill_type="solid", fgColor="4F81BD")

        for i, row in ranked_df.head(5).iterrows():
            ws.cell(row=start + i + 1, column=1).value = row["Rank"]

            ws.cell(row=start + i + 1, column=2).value = row["Strategy"]

            ws.cell(row=start + i + 1, column=3).value = round(row["Overall Score"], 2)

            ws.cell(row=start + i + 1, column=4).value = row["Grade"]

    #######################################################################

    def metric_summary(self, ws, ranked_df):
        row = 20

        ws[f"A{row}"] = "Average Score"

        ws[f"B{row}"] = round(ranked_df["Overall Score"].mean(), 2)

        row += 1

        ws[f"A{row}"] = "Highest Score"

        ws[f"B{row}"] = round(ranked_df["Overall Score"].max(), 2)

        row += 1

        ws[f"A{row}"] = "Lowest Score"

        ws[f"B{row}"] = round(ranked_df["Overall Score"].min(), 2)

        row += 1

        ws[f"A{row}"] = "Average Grade"

        ws[f"B{row}"] = "-"
