import openpyxl

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment

from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment

from openpyxl.workbook.properties import CalcProperties

from openpyxl.formatting.rule import ColorScaleRule

from openpyxl.utils import get_column_letter

from config import OUTPUT_FILE


class ExcelWriterEngine:

    def __init__(self):

        self.writer = None

    ###########################################################################
    # OPEN WRITER
    ###########################################################################

    def open(self):

        self.writer = openpyxl.Workbook()

        self.writer.remove(
            self.writer.active
        )

    ###########################################################################
    # WRITE DATAFRAME
    ###########################################################################

    def write_sheet(
        self,
        dataframe,
        sheet_name
    ):

        ws = self.writer.create_sheet(
            title=sheet_name
        )

        ws.append(
            list(dataframe.columns)
        )

        for row in dataframe.itertuples(index=False):

            ws.append(list(row))

        self.format_sheet(ws)

    ###########################################################################
    # MAIN FORMATTER
    ###########################################################################

    def format_sheet(self, ws):

        self.header_style(ws)

        self.freeze(ws)

        self.auto_filter(ws)

        self.auto_width(ws)

        self.align(ws)

        self.borders(ws)

        self.highlight_numeric(ws)

    ###########################################################################
    # HEADER
    ###########################################################################

    def header_style(self, ws):

        fill = PatternFill(

            fill_type="solid",

            fgColor="1F4E78"

        )

        font = Font(

            bold=True,

            color="FFFFFF",

            size=11

        )

        for cell in ws[1]:

            cell.fill = fill

            cell.font = font

            cell.alignment = Alignment(

                horizontal="center",

                vertical="center"

            )

    ###########################################################################
    # FREEZE PANE
    ###########################################################################

    def freeze(self, ws):

        ws.freeze_panes = "A2"

    ###########################################################################
    # FILTER
    ###########################################################################

    def auto_filter(self, ws):

        ws.auto_filter.ref = ws.dimensions

    ###########################################################################
    # AUTO WIDTH
    ###########################################################################

    def auto_width(self, ws):

        for column_cells in ws.columns:

            length = max(

                len(str(cell.value))

                if cell.value is not None

                else 0

                for cell in column_cells

            )

            column = get_column_letter(

                column_cells[0].column

            )

            ws.column_dimensions[column].width = min(

                length + 4,

                40

            )

    ###########################################################################
    # SHEET ORDER
    ###########################################################################

    def reorder_sheets(self):

        preferred = [

            "Dashboard",

            "Executive Summary",

            "Executive Insights",

            "Strategy Ranking",

            "Metric Winners",

            "Metric Leaders",

            "Score Breakdown",

            "Recommendations",

            "Deployment Guide",

            "Strengths & Weaknesses",

            "Top 10 Stocks"

        ]

        ordered = []

        for name in preferred:

            if name in self.writer.sheetnames:

                ordered.append(

                    self.writer[name]

                )

        for sheet in self.writer.worksheets:

            if sheet not in ordered:

                ordered.append(sheet)

        self.writer._sheets = ordered

    ###########################################################################
    # DASHBOARD ACTIVE
    ###########################################################################

    def activate_dashboard(self):

        if "Dashboard" in self.writer.sheetnames:

            self.writer.active = self.writer["Dashboard"]

    ###########################################################################
    # WORKBOOK PROPERTIES
    ###########################################################################

    def workbook_properties(self):

        self.writer.properties.creator = "Strategy Comparison Engine"

        self.writer.properties.title = "Strategy Comparison Report"

        self.writer.properties.subject = "Trading Strategy Analysis"

        self.writer.properties.description = (

            "Automatically generated comparison report."

        )

        self.writer.properties.category = "Trading"

        self.writer.properties.keywords = (

            "Strategy Comparison"

        )

        self.writer.calculation = CalcProperties(

            calcMode="auto"

        )

    ###########################################################################
    # NAVIGATION
    ###########################################################################

    def navigation(self):

        if "Dashboard" not in self.writer.sheetnames:

            return

        dashboard = self.writer["Dashboard"]

        dashboard["J2"] = "Navigation"

        dashboard["J2"].font = Font(

            bold=True,

            color="FFFFFF"

        )

        dashboard["J2"].fill = PatternFill(

            fill_type="solid",

            fgColor="4F81BD"

        )

        dashboard["J2"].alignment = Alignment(

            horizontal="center"

        )

        row = 3

        for sheet in self.writer.sheetnames:

            if sheet == "Dashboard":

                continue

            cell = dashboard[f"J{row}"]

            cell.value = sheet

            cell.hyperlink = f"#'{sheet}'!A1"

            cell.style = "Hyperlink"

            row += 1

    ###########################################################################
    # BACK TO DASHBOARD
    ###########################################################################

    def back_to_dashboard(self):

        for sheet in self.writer.sheetnames:

            if sheet == "Dashboard":

                continue

            ws = self.writer[sheet]

            ws["A1"] = "← Dashboard"

            ws["A1"].hyperlink = "#Dashboard!A1"

            ws["A1"].style = "Hyperlink"

    ###########################################################################
    # ALIGNMENT
    ###########################################################################

    def align(self, ws):

        for row in ws.iter_rows():

            for cell in row:

                cell.alignment = Alignment(

                    horizontal="center",

                    vertical="center"

                )

    ###########################################################################
    # BORDER
    ###########################################################################

    def borders(self, ws):

        thin = Side(

            style="thin",

            color="D9D9D9"

        )

        border = Border(

            left=thin,

            right=thin,

            top=thin,

            bottom=thin

        )

        for row in ws.iter_rows():

            for cell in row:

                cell.border = border

    ###########################################################################
    # CONDITIONAL FORMATTING
    ###########################################################################

    def highlight_numeric(self, ws):

        if ws.max_row <= 1:
            return

        for col in ws.iter_cols():

            numeric = True
            has_numeric = False

            for cell in col[1:]:

                if cell.value is None:
                    continue

                if isinstance(cell.value, (int, float)):
                    has_numeric = True
                else:
                    numeric = False
                    break

            if not numeric or not has_numeric:
                continue

            if ws.max_row < 2:
                continue

            letter = get_column_letter(col[0].column)

            rng = f"{letter}2:{letter}{ws.max_row}"

            ws.conditional_formatting.add(

                rng,

                ColorScaleRule(

                    start_type="min",
                    start_color="F8696B",

                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",

                    end_type="max",
                    end_color="63BE7B"

                )

            )

    ###########################################################################
    # SAVE
    ###########################################################################

    def save(self):

        self.reorder_sheets()

        self.activate_dashboard()

        self.navigation()

        self.back_to_dashboard()

        self.workbook_properties()

        self.writer.save(

            OUTPUT_FILE

        )