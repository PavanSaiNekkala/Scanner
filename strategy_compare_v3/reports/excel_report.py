"""
============================================================
Institutional Strategy Comparison Engine V3
File : reports/excel_report.py

Excel Report Generator

Author : Pavan Sai
============================================================
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font

from core.logger import get_logger

logger = get_logger(__name__)


class ExcelReport:
    """
    Generates a professional Excel workbook
    containing all analysis results.
    """

    def __init__(
        self,
        output_directory: str = "outputs/excel",
        filename: str = "Institutional_Report.xlsx",
    ):
        self.output_directory = Path(output_directory)

        self.output_directory.mkdir(parents=True, exist_ok=True)

        self.output_file = self.output_directory / filename

    # --------------------------------------------------

    @staticmethod
    def _format_sheet(ws):
        header = Font(bold=True)

        for cell in ws[1]:
            cell.font = header

        ws.freeze_panes = "A2"

        for column in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column
            )

            ws.column_dimensions[column[0].column_letter].width = min(length + 3, 40)

    # --------------------------------------------------

    def generate(self, reports: Dict[str, pd.DataFrame]) -> str:
        logger.info("Generating Excel Report...")

        with pd.ExcelWriter(self.output_file, engine="openpyxl") as writer:
            for sheet, dataframe in reports.items():
                dataframe.to_excel(writer, sheet_name=sheet[:31], index=False)

        workbook = load_workbook(self.output_file)

        for sheet in workbook.sheetnames:
            self._format_sheet(workbook[sheet])

        workbook.save(self.output_file)

        logger.info("Excel report saved to %s", self.output_file)

        return str(self.output_file)


if __name__ == "__main__":
    print("Import inside report_engine.py")
