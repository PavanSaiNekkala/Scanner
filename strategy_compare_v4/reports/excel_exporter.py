"""
=============================================================
Institutional Strategy Comparison Platform V4

Module
------
reports/excel_exporter.py

Purpose
-------
Export institutional reports into a professionally
formatted Excel workbook.

Features
--------
• Multiple Worksheets
• Auto Column Width
• Freeze Panes
• Auto Filters
• Number Formatting
• Conditional Formatting
• Institution Ready Layout

=============================================================
"""

from __future__ import annotations

import time
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import (
    ColorScaleRule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import (
    get_column_letter,
)

from strategy_compare_v4.config.constants import (
    RECOMMENDATION,
)
from strategy_compare_v4.reports.final_columns import FINAL_COLUMNS
from strategy_compare_v4.utils.logger import (
    banner,
    get_logger,
)

logger = get_logger(__name__)


# ============================================================
# Excel Export Engine
# ============================================================


class ExcelExporter:
    """
    Institutional Excel
    Export Engine.
    """

    from pathlib import Path

    def __init__(
        self,
        output_file: str | Path,
    ):
        self.filename = Path(output_file)
        self.output = output_file

        self.writer: pd.ExcelWriter | None = None

        self.workbook = None

        self.execution_time: float = 0.0

        self.diagnostic_report: dict = {}

    # ---------------------------------------------------------
    # Create Workbook
    # ---------------------------------------------------------

    def create_workbook(
        self,
    ):
        """
        Create workbook.
        """

        banner(
            logger,
            "Creating Excel Workbook",
        )

        self.writer = pd.ExcelWriter(
            self.output,
            engine="openpyxl",
        )

        logger.info("Workbook created.")

        logger.info(
            "Output : %s",
            self.output,
        )

        return self

    # ---------------------------------------------------------
    # Add Worksheet
    # ---------------------------------------------------------

    def add_sheet(
        self,
        dataframe: pd.DataFrame,
        sheet_name: str,
    ):
        """
        Export dataframe to
        workbook.
        """

        if self.writer is None:
            raise RuntimeError("Workbook has not been created.")

        dataframe.to_excel(
            self.writer,
            sheet_name=sheet_name,
            index=False,
        )

        logger.info(
            "Sheet Added : %-25s Rows=%d",
            sheet_name,
            len(
                dataframe,
            ),
        )

        return self

    # ---------------------------------------------------------
    # Save Workbook
    # ---------------------------------------------------------

    def save(
        self,
    ):
        """
        Save workbook to disk.
        """

        if self.writer is None:
            raise RuntimeError("Workbook has not been created.")

        self.writer.close()

        logger.info("Workbook saved.")

        return self

    # ---------------------------------------------------------
    # Load Workbook
    # ---------------------------------------------------------

    def load_workbook(
        self,
    ):
        """
        Load workbook for
        post-processing and
        formatting.
        """

        banner(
            logger,
            "Loading Workbook",
        )

        self.workbook = openpyxl.load_workbook(
            self.output,
        )

        logger.info("Workbook loaded.")

        logger.info(
            "Worksheets : %d",
            len(
                self.workbook.sheetnames,
            ),
        )

        return self

    # ---------------------------------------------------------
    # Style Headers
    # ---------------------------------------------------------

    def style_headers(
        self,
    ):
        """
        Apply institutional
        header formatting to
        every worksheet.
        """

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        thin = Side(
            style="thin",
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        for worksheet in self.workbook.worksheets:
            for cell in worksheet[1]:
                cell.fill = header_fill

                cell.font = header_font

                cell.border = border

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

        logger.info("Headers formatted.")

        return self

    # ---------------------------------------------------------
    # Freeze Headers
    # ---------------------------------------------------------

    def freeze_headers(
        self,
    ):
        """
        Freeze first row
        across all worksheets.
        """

        for worksheet in self.workbook.worksheets:
            worksheet.freeze_panes = "A2"

        logger.info("Freeze panes applied.")

        return self

    # ---------------------------------------------------------
    # Auto Filter
    # ---------------------------------------------------------

    def auto_filter(
        self,
    ):
        """
        Enable worksheet
        auto-filters.
        """

        for worksheet in self.workbook.worksheets:
            worksheet.auto_filter.ref = worksheet.dimensions

        logger.info("Auto filters enabled.")

        return self

    # ---------------------------------------------------------
    # Auto Column Width
    # ---------------------------------------------------------

    def auto_width(
        self,
    ):
        """
        Automatically adjust
        column widths.
        """

        for worksheet in self.workbook.worksheets:
            for column in worksheet.columns:
                max_length = 0

                column_letter = get_column_letter(
                    column[0].column,
                )

                for cell in column:
                    try:
                        value = (
                            ""
                            if cell.value is None
                            else str(
                                cell.value,
                            )
                        )

                        max_length = max(
                            max_length,
                            len(
                                value,
                            ),
                        )

                    except Exception:
                        continue

                worksheet.column_dimensions[column_letter].width = min(
                    max_length + 3,
                    40,
                )

        logger.info("Column widths adjusted.")

        return self

    # ---------------------------------------------------------
    # Number Formatting
    # ---------------------------------------------------------

    def number_format(
        self,
    ):
        """
        Apply institutional
        numeric formatting.
        """

        for worksheet in self.workbook.worksheets:
            for row in worksheet.iter_rows(
                min_row=2,
            ):
                for cell in row:
                    if isinstance(
                        cell.value,
                        (
                            int,
                            float,
                        ),
                    ):
                        cell.number_format = "0.00"

        logger.info("Number formatting applied.")

        return self

    # ---------------------------------------------------------
    # Apply Color Scale
    # ---------------------------------------------------------

    def apply_color_scale(
        self,
    ):
        """
        Apply conditional
        color-scale formatting
        to numeric columns.
        """

        rule = ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        )

        formatted_columns = 0

        for worksheet in self.workbook.worksheets:
            if worksheet.max_row <= 2:
                continue

            for column in range(
                2,
                worksheet.max_column + 1,
            ):
                worksheet.conditional_formatting.add(
                    (
                        f"{get_column_letter(column)}2:"
                        f"{get_column_letter(column)}"
                        f"{worksheet.max_row}"
                    ),
                    rule,
                )

                formatted_columns += 1

        logger.info(
            "Conditional formatting applied to %d column(s).",
            formatted_columns,
        )

        return self

    # ---------------------------------------------------------
    # Alternate Row Colors
    # ---------------------------------------------------------

    def alternate_row_colors(
        self,
    ):
        """
        Apply alternating row
        background colors.
        """

        alternate_fill = PatternFill(
            fill_type="solid",
            fgColor="F7F7F7",
        )

        for worksheet in self.workbook.worksheets:
            for row in range(
                2,
                worksheet.max_row + 1,
            ):
                if row % 2 != 0:
                    continue

                for column in range(
                    1,
                    worksheet.max_column + 1,
                ):
                    worksheet.cell(
                        row=row,
                        column=column,
                    ).fill = alternate_fill

        logger.info("Alternate row formatting applied.")

        return self

    # ---------------------------------------------------------
    # Highlight Recommendations
    # ---------------------------------------------------------

    def highlight_recommendations(
        self,
    ):
        """
        Highlight recommendation
        values using institutional
        colors.
        """

        recommendation_colors = {
            "Strong Buy": "00B050",
            "Buy": "92D050",
            "Watch": "FFD966",
            "Improve": "F4B183",
            "Avoid": "FF9999",
            "Reject": "FF0000",
        }

        highlighted = 0

        for worksheet in self.workbook.worksheets:
            headers = [cell.value for cell in worksheet[1]]

            if RECOMMENDATION not in headers:
                continue

            recommendation_column = (
                headers.index(
                    RECOMMENDATION,
                )
                + 1
            )

            for row in range(
                2,
                worksheet.max_row + 1,
            ):
                cell = worksheet.cell(
                    row=row,
                    column=recommendation_column,
                )

                if cell.value not in recommendation_colors:
                    continue

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=(recommendation_colors[cell.value]),
                )

                cell.font = Font(
                    bold=True,
                    color="FFFFFF",
                )

                highlighted += 1

        logger.info(
            "Highlighted %d recommendation cells.",
            highlighted,
        )

        return self

    # ---------------------------------------------------------
    # Summary Sheet
    # ---------------------------------------------------------

    def add_summary_sheet(
        self,
    ):
        """
        Create workbook
        summary sheet.
        """

        if "Summary" in self.workbook.sheetnames:
            del self.workbook["Summary"]

        worksheet = self.workbook.create_sheet(
            "Summary",
            0,
        )

        worksheet["A1"] = "Institutional Strategy Comparison Report"

        worksheet["A1"].font = Font(
            bold=True,
            size=16,
        )

        worksheet["A3"] = "Workbook"

        worksheet["B3"] = Path(
            self.output,
        ).name

        worksheet["A4"] = "Worksheets"

        worksheet["B4"] = (
            len(
                self.workbook.sheetnames,
            )
            - 1
        )

        worksheet["A5"] = "Generated"

        worksheet["B5"] = pd.Timestamp.now()

        logger.info("Summary sheet created.")

        return self

    # ---------------------------------------------------------
    # Workbook Properties
    # ---------------------------------------------------------

    def workbook_properties(
        self,
    ):
        """
        Configure workbook
        metadata.
        """

        properties = self.workbook.properties

        properties.creator = "Strategy Comparison Engine"

        properties.title = "Institutional Strategy Report"

        properties.subject = "Strategy Analytics"

        properties.category = "Trading Analytics"

        logger.info("Workbook metadata configured.")

        return self

    # ---------------------------------------------------------
    # Finalize Workbook
    # ---------------------------------------------------------

    def finalize(
        self,
    ):
        """
        Save the formatted
        workbook.
        """

        self.workbook.save(
            self.output,
        )

        self.diagnostic_report = {
            "output_file": self.output,
            "worksheets": len(
                self.workbook.sheetnames,
            ),
            "status": "Success",
        }

        logger.info("Workbook finalized.")

        logger.info(
            "Saved As : %s",
            self.output,
        )

        return self

    # ---------------------------------------------------------
    # Execution Report
    # ---------------------------------------------------------

    def execution_report(
        self,
    ):
        """
        Log execution
        statistics.
        """

        banner(
            logger,
            "Excel Export Report",
        )

        logger.info(
            "Execution Time : %.3f seconds",
            self.execution_time,
        )

        logger.info(
            "Workbook       : %s",
            self.output,
        )

        logger.info(
            "Worksheets     : %d",
            len(
                self.workbook.sheetnames,
            ),
        )

        return self

    # ---------------------------------------------------------
    # Run Formatting Pipeline
    # ---------------------------------------------------------

    def run(
        self,
    ):
        """
        Execute the complete
        formatting pipeline.
        """

        start = time.perf_counter()

        try:
            (
                self.load_workbook()
                .style_headers()
                .freeze_headers()
                .auto_filter()
                .auto_width()
                .number_format()
                .apply_color_scale()
                .alternate_row_colors()
                .highlight_recommendations()
                .add_summary_sheet()
                .workbook_properties()
                .finalize()
            )

        except Exception:
            logger.exception("Excel export failed.")

            raise

        finally:
            self.execution_time = time.perf_counter() - start

            self.execution_report()

        return self


# ============================================================
# Convenience Function
# ============================================================


def export_excel(
    output_file: str,
    sheets: dict[str, pd.DataFrame],
):
    """
    Export multiple
    dataframes to a
    professionally formatted
    Excel workbook.

    Parameters
    ----------
    output_file : str
        Destination workbook.

    sheets : dict
        Dictionary of
        worksheet name →
        dataframe.
    """

    exporter = ExcelExporter(
        output_file,
    )

    exporter.create_workbook()

    for sheet_name, dataframe in sheets.items():
        if not isinstance(
            dataframe,
            pd.DataFrame,
        ):
            logger.warning(
                "Skipped '%s' (not a DataFrame).",
                sheet_name,
            )

            continue

        if sheet_name == "Comparison":
            dataframe = dataframe[
                [column for column in FINAL_COLUMNS if column in dataframe.columns]
            ]

        exporter.add_sheet(
            dataframe,
            sheet_name,
        )

    exporter.save()

    exporter.run()

    logger.info("Excel export completed successfully.")

    return exporter


# ============================================================
# Main
# ============================================================

if __name__ == "__main__":
    logger.info("Import export_excel() into the reporting pipeline.")
